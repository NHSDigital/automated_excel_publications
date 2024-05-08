import config
import dateutil
from typing import Tuple, List
import openpyxl
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows

# region UTILITIES

def get_list_of_months() -> list:
    """
    Using the config file to fetch the current publication month and number of months included,
    loops backwards from current month to get a list of months included in the publication in the
    format required by the table 1 columns (MMM-YY)

    Args: None

    Returns:
        list: list_of_months
    """
    starting_month = config.get_report_month()
    list_of_months = [starting_month]
    delta = dateutil.relativedelta.relativedelta(months=-1)

    for i in range(config.get_number_of_months() - 1):
        month_datetime = list_of_months[-1]
        list_of_months += [month_datetime + delta]
    list_of_months = [i.strftime("%b-%y") for i in list_of_months]
    return list_of_months


def filter_df_to_report_month(df: pd.DataFrame) -> pd.DataFrame:
    """
    Filters ingested data down to the month of the publication

    Args:
        df (pd.DataFrame): The ingested dataframe

    Returns:
        pd.DataFrame: The filtered dataframe
    """
    report_month = config.get_report_month()
    df = df[df.appt_date != "ALL"]
    pd.options.mode.chained_assignment = None  # default='warn'
    df["datetime_date"] = pd.to_datetime(df["appt_date"], format="%Y-%m-%d")
    df = df[df["datetime_date"].dt.month == report_month.month]
    df = df[df["datetime_date"].dt.year == report_month.year]
    df.drop(["datetime_date"], axis=1, inplace=True)
    return df


def find_cell_in_column(ws: openpyxl.worksheet, tag: str, column: str) -> Tuple:
    """
    Finds a cell with a tag in a particular column

    Args:
        ws (openpyxl.worksheet): The worksheet
        tag (str): The tag in the ell
        column (str): The column you're looking in, A or B etc.

    Returns:
        Tuple: The index of the cell containing the tag
    """
    max_search = 1000
    iter = 0
    while iter <= max_search:
        for cell in ws[column]:
            iter += 1
            if cell.value == tag:
                return openpyxl.utils.cell.coordinate_to_tuple(cell.coordinate)
    return None


def write_df_from_start_cell(
    start_cell: Tuple, end_cell: Tuple, ws: openpyxl.worksheet, df: pd.DataFrame
) -> None:
    """
    Given a pandas dataframe and a worksheet, writes that dataframe to that worksheet.
    Starts at the cell with the <start> tag. Any empty rows after the last df row are written,
    and up to the <end> cell row, are deleted.

    Args:
        start_cell (Tuple): Cell to start the data in
        end_cell (Tuple): Cell to delete blank rows up to
        ws (openpyxl.worksheet): The worksheet to write to
        df (pd.DataFrame): The data to write
    """
    rows_to_write = dataframe_to_rows(df, index=False, header=False)
    loc = list(start_cell)
    for row in rows_to_write:
        for cell in row:
            ws.cell(row=loc[0], column=loc[1]).value = cell
            loc[1] += 1
        loc[0] += 1
        loc[1] = start_cell[1]
    clear_empty_rows(ws=ws, last_written_row=loc[0], end_cell=end_cell)


def clear_empty_rows(
    ws: openpyxl.worksheet, last_written_row: int, end_cell: Tuple
) -> None:
    """
    Deletes blank / empty rows from the area where data is written to.
    This is a workaround for a problem with working from templates; how do we size the template without knowing how many rows the df will have?
    To get round this, we insert many many blank rows, and then delete the unused ones.

    Args:
        ws (openpyxl.worksheet): the worksheet to clear empty rows from
        last_written_row (int): The last row to which we wrote data
        end_cell (Tuple): The location of the <end> tag
    """
    number_to_delete = end_cell[0] - last_written_row
    ws.delete_rows(last_written_row + 1, number_to_delete)


def write_table_to_sheet(
    wb: openpyxl.Workbook, table_data: pd.DataFrame, sheet_name: str
) -> openpyxl.Workbook:
    """
    Given some data, a workbook, and a sheet name; writes that data to the chosen sheet

    Args:
        wb (openpyxl.Workbook): The workbook
        table_data (pd.DataFrame): The data to write
        sheet_name (str): The sheet to write to

    Returns:
        openpyxl.Workbook: The workbook, with the data written
    """
    start_cell = find_cell_by_tag(wb, sheet_name, "<start>")
    end_cell = find_cell_by_tag(wb, sheet_name, "<end>")

    ws = wb[sheet_name]
    write_df_from_start_cell(
        start_cell=start_cell, end_cell=end_cell, ws=ws, df=table_data
    )
    return wb


def make_and_write_easy_a(wb: openpyxl.Workbook) -> openpyxl.Workbook:
    """
    Writes sheet to workbook in the 'easy' example

    Args:
        wb (openpyxl.Workbook): The workbook which has been loaded from template

    Returns:
        openpyxl.Workbook: The same workbook, but with the sheet written in
    """    

    # Load the dataframe in from the datafile
    df = config.get_easy_a_data()

    # Make sure that the column order matches the column order in the template
    df = df[["weekday", "appt_date", "total", "Attended", "DNA", "Unknown"]]

    # Write the table to the workbook sheet
    wb = write_table_to_sheet(wb=wb, table_data=df, sheet_name="Easy A")
    return wb


def make_and_write_easy_b(wb: openpyxl.Workbook) -> openpyxl.Workbook:
    """
    Writes sheet to workbook in the 'easy' example

    Args:
        wb (openpyxl.Workbook): The workbook which has been loaded from template

    Returns:
        openpyxl.Workbook: The same workbook, but with the sheet written in
    """    
    # Load the dataframe in from the datafile
    df = config.get_easy_a_data()

    # Make sure that the column order matches the column order in the template
    df = df[["weekday", "appt_date", "total", "Attended", "DNA", "Unknown"]]

    # Write the table to the workbook sheet
    wb = write_table_to_sheet(wb=wb, table_data=df, sheet_name="Easy B")
    return wb


def make_and_write_2a(wb: openpyxl.Workbook) -> openpyxl.Workbook:
    """
    Writes sheet '2a' to the workbook. Loads in data and does some basic organising first. 

    Args:
        wb (openpyxl.Workbook): The workbook loaded from template

    Returns:
        openpyxl.Workbook: The workbook, with the sheet written.
    """    
    df = config.get_appointments_data()
    df = df[df["breakdown"] == "by_status_by_date"]
    df = df[["appt_date", "appt_status", "appt_count"]]
    df = df.pivot_table(
        index="appt_date", columns="appt_status", values="appt_count"
    ).fillna(0)

    df.index = pd.to_datetime(df.index).strftime("%d/%b/%y")

    df["total"] = df.sum(axis=1)
    df["weekday"] = pd.to_datetime(df.index).strftime("%a")
    df = df.reset_index(level=0)
    df = df[["weekday", "appt_date", "total", "Attended", "DNA", "Unknown"]]

    wb = write_table_to_sheet(wb=wb, table_data=df, sheet_name="Table 2a")
    return wb


def make_and_write_2b(wb: openpyxl.Workbook) -> openpyxl.Workbook:
    """
    Writes sheet '2b' to the workbook. Loads in data and does some basic organising first. 

    Args:
        wb (openpyxl.Workbook): The workbook loaded from template

    Returns:
        openpyxl.Workbook: The workbook, with the sheet written.
    """    
    df = config.get_appointments_data()
    df = df[df["breakdown"] == "by_hcp_type_by_date"]
    df = df[["appt_date", "hcp_type", "appt_count"]]
    df = df.pivot_table(
        index="appt_date", columns="hcp_type", values="appt_count"
    ).fillna(0)

    df.index = pd.to_datetime(df.index).strftime("%d/%b/%y")

    df["total"] = df.sum(axis=1)
    df["weekday"] = pd.to_datetime(df.index).strftime("%a")
    df = df.reset_index(level=0)
    df = df[["weekday", "appt_date", "total", "GP", "Other Practice Staff", "Unknown"]]

    wb = write_table_to_sheet(wb=wb, sheet_name="Table 2b", table_data=df)

    return wb


def make_and_write_2c(wb: openpyxl.Workbook) -> openpyxl.Workbook:
    """
    Writes sheet '2c' to the workbook. Loads in data and does some basic organising first. 

    Args:
        wb (openpyxl.Workbook): The workbook loaded from template

    Returns:
        openpyxl.Workbook: The workbook, with the sheet written.
    """    

    df = config.get_appointments_data()
    df = df[df["breakdown"] == "by_appt_mode_by_date"]
    df = df[["appt_date", "appt_mode", "appt_count"]]
    df = df.pivot_table(
        index="appt_date", columns="appt_mode", values="appt_count"
    ).fillna(0)
    df.index = pd.to_datetime(df.index).strftime("%d/%b/%y")

    df["total"] = df.sum(axis=1)
    df["weekday"] = pd.to_datetime(df.index).strftime("%a")
    df = df.reset_index(level=0)
    df = df[
        [
            "weekday",
            "appt_date",
            "total",
            "Face-to-Face",
            "Home Visit",
            "Telephone",
            "Video/Online",
            "Unknown",
        ]
    ]
    wb = write_table_to_sheet(wb=wb, sheet_name="Table 2c", table_data=df)
    return wb


def make_and_write_2d(wb: openpyxl.Workbook) -> openpyxl.Workbook:
    """
    Writes sheet '2d' to the workbook. Loads in data and does some basic organising first. 

    Args:
        wb (openpyxl.Workbook): The workbook loaded from template

    Returns:
        openpyxl.Workbook: The workbook, with the sheet written.
    """    

    df = config.get_appointments_data()
    df = df[df["breakdown"] == "by_time_between_booking_and_appt_by_date"]
    df = df[["appt_date", "time_between_booking_and_appt", "appt_count"]]
    df = df.pivot_table(
        index="appt_date", columns="time_between_booking_and_appt", values="appt_count"
    ).fillna(0)

    df.index = pd.to_datetime(df.index).strftime("%d/%b/%y")

    df["total"] = df.sum(axis=1)
    df["weekday"] = pd.to_datetime(df.index).strftime("%a")
    df = df.reset_index(level=0)
    df = df[
        [
            "weekday",
            "appt_date",
            "total",
            "Same Day",
            "1 Day",
            "2 to 7 Days",
            "8 to 14 Days",
            "15 to 21 Days",
            "22 to 28 Days",
            "More than 28 Days",
            "Unknown / Data Quality",
        ]
    ]
    wb = write_table_to_sheet(wb=wb, sheet_name="Table 2d", table_data=df)
    return wb


def combine_appts_with_practices(
    breakdowns_set: set, appointments_pivot: str, pivoted_column_list: List[str]
) -> pd.DataFrame:
    """
    Summary: For some sheets, we want data from two different sources; 'appointments' and 'practices'.
    Ultimately we want to display the data from these indexed by geography. This function loads these two sources in and
    joins them in the desired way. Later functions can then filter down to the desired columns. 

    Args:
        breakdowns_set (set): Set of relevant breakdowns to include
        appointments_pivot (str): Column of categorical data: the values of this column will correspond to the column headings in the Excel
        pivoted_column_list (List[str]): The ordered list of column headings: these ought to match the order of headings in the Excel

    Returns:
        pd.DataFrame: The combined dataframe, joined by geography, and sorted according to the size of geographic region. 
    """

    # Ingest the data
    df_appts = config.get_appointments_data()
    df_practices = config.get_practices_data()

    # We will want to sort our geographies; the following dict is for that purpose. 
    custom_dict = {'National': 0, 'Region': 1, 'STP': 2, 'CCG': 3}
    df_practices['rank'] = df_practices['geog_type'].map(custom_dict)

    df_practices.sort_values(by=['rank', 'geog_code'], ascending = [True, True], inplace=True)
    #Now sorted, so we can drop the rank
    df_practices.drop(columns = ['rank'], inplace=True)

    # Prepare the practices data
    df_practices = df_practices[
        [
            "geog_type",
            "geog_ons_code",
            "count_of_open_practice",
            "count_of_included_practice",
        ]
    ]
    df_practices = df_practices.set_index("geog_ons_code")

    # Prepare the appointments data
    df_appts = df_appts[df_appts["breakdown"].isin(breakdowns_set)]
    df_appts = df_appts[
        [appointments_pivot, "geog_name", "geog_code", "geog_ons_code", "appt_count"]
    ]
    df_geogs = df_appts[["geog_name", "geog_code", "geog_ons_code"]].set_index(
        "geog_ons_code"
    )
    df_appts = df_appts.pivot_table(
        index="geog_ons_code", columns=appointments_pivot, values="appt_count"
    ).fillna(0)
    df_appts["total"] = df_appts.sum(axis=1) #! Again, total in DAE?
    df_appts = df_appts.reset_index(level=0)
    df_appts = df_appts.set_index("geog_ons_code")
    df_appts = df_appts.join(df_geogs, how="inner").drop_duplicates()

    # Combine the appointments and practices data
    df_combined = df_practices.join(df_appts, how="inner")
    df_combined.reset_index(inplace=True)
    df_combined = df_combined.drop_duplicates()
    column_list = [
        "geog_type",
        "geog_code",
        "geog_ons_code",
        "geog_name",
        "count_of_open_practice",
        "count_of_included_practice",
        "total",
    ] + pivoted_column_list
    df_combined = df_combined[column_list]

    return df_combined


def make_and_write_3a(wb: openpyxl.Workbook) -> openpyxl.Workbook:
    breakdowns_set = {
        "national_count_by_appt_status",
        "by_ccg_and_appt_status",
        "by_stp_and_appt_status",
        "by_region_and_appt_status",
    }
    appointments_pivot = "appt_status"
    pivoted_column_list = [
        "Attended",
        "DNA",
        "Unknown",
    ]
    df = combine_appts_with_practices(
        breakdowns_set=breakdowns_set,
        appointments_pivot=appointments_pivot,
        pivoted_column_list=pivoted_column_list,
    )

    wb = write_table_to_sheet(wb=wb, sheet_name="Table 3a", table_data=df)
    return wb


def make_and_write_3b(wb: openpyxl.Workbook) -> openpyxl.Workbook:
    """
    Writes sheet '3b' to the workbook. Loads in data and does some basic organising first. 

    Args:
        wb (openpyxl.Workbook): The workbook loaded from template

    Returns:
        openpyxl.Workbook: The workbook, with the sheet written.
    """    

    breakdowns_set = {
        "national_count_by_hcp_type",
        "by_ccg_and_hcp_type",
        "by_stp_and_hcp_type",
        "by_region_and_hcp_type",
    }

    appointments_pivot = "hcp_type"
    pivoted_column_list = [
        "GP",
        "Other Practice Staff",
        "Unknown",
    ]
    df = combine_appts_with_practices(
        breakdowns_set=breakdowns_set,
        appointments_pivot=appointments_pivot,
        pivoted_column_list=pivoted_column_list,
    )

    wb = write_table_to_sheet(wb=wb, sheet_name="Table 3b", table_data=df)
    return wb


def make_and_write_3c(wb: openpyxl.Workbook) -> openpyxl.Workbook:
    """
    Writes sheet '3c' to the workbook. Loads in data and does some basic organising first. 

    Args:
        wb (openpyxl.Workbook): The workbook loaded from template

    Returns:
        openpyxl.Workbook: The workbook, with the sheet written.
    """    
    breakdowns_set = {
        "national_count_by_appt_mode",
        "by_ccg_and_appt_mode",
        "by_stp_and_appt_mode",
        "by_region_and_appt_mode",
    }

    appointments_pivot = "appt_mode"
    pivoted_column_list = [
        "Face-to-Face",
        "Home Visit",
        "Telephone",
        "Video/Online",
        "Unknown",
    ]
    df = combine_appts_with_practices(
        breakdowns_set=breakdowns_set,
        appointments_pivot=appointments_pivot,
        pivoted_column_list=pivoted_column_list,
    )

    wb = write_table_to_sheet(wb=wb, sheet_name="Table 3c", table_data=df)
    return wb


def make_and_write_3d(wb: openpyxl.Workbook) -> openpyxl.Workbook:
    """
    Writes sheet '3d' to the workbook. Loads in data and does some basic organising first. 

    Args:
        wb (openpyxl.Workbook): The workbook loaded from template

    Returns:
        openpyxl.Workbook: The workbook, with the sheet written.
    """    
    breakdowns_set = {
        "national_count_by_time_between_booking_and_appt",
        "by_ccg_and_time_between_booking_and_appt",
        "by_stp_and_time_between_booking_and_appt",
        "by_region_and_time_between_booking_and_appt",
    }

    appointments_pivot = "time_between_booking_and_appt"
    pivoted_column_list = [
        "Same Day",
        "1 Day",
        "2 to 7 Days",
        "8 to 14 Days",
        "15 to 21 Days",
        "22 to 28 Days",
        "More than 28 Days",
        "Unknown / Data Quality",
    ]
    df = combine_appts_with_practices(
        breakdowns_set=breakdowns_set,
        appointments_pivot=appointments_pivot,
        pivoted_column_list=pivoted_column_list,
    )

    wb = write_table_to_sheet(wb=wb, sheet_name="Table 3d", table_data=df)
    return wb


def make_and_write_3e(wb: openpyxl.Workbook) -> openpyxl.Workbook:
    """
    Writes sheet '3e' to the workbook. Loads in data and does some basic organising first. 

    Args:
        wb (openpyxl.Workbook): The workbook loaded from template

    Returns:
        openpyxl.Workbook: The workbook, with the sheet written.
    """    

    # Ingest the data
    df_appts = config.get_appointments_data()

    # Prepare the practices data
    df_list_size = df_list_size[
        ["geog_ons_code", "open_practice_count", "included_practice_count"]
    ]
    df_list_size = df_list_size.set_index("geog_ons_code")

    # Prepare the appointments data
    df_appts = df_appts[df_appts["breakdown"] == "by_ccg_and_appt_mode"]
    df_appts = df_appts[
        [
            "appt_mode",
            "geog_name",
            "geog_code",
            "geog_ons_code",
            "appt_count",
        ]
    ]
    df_geogs = df_appts[["geog_name", "geog_code", "geog_ons_code"]].set_index(
        "geog_ons_code"
    )
    df_appts = df_appts.pivot_table(
        index="geog_ons_code",
        columns="appt_mode",
        values="appt_count",
    ).fillna(0)
    df_appts["total"] = df_appts.sum(axis=1)
    df_appts = df_appts.reset_index(level=0)
    df_appts = df_appts.set_index("geog_ons_code")
    df_appts = df_appts.join(df_geogs, how="inner")

    # Combine the appointments and practices data
    df_combined = df_list_size.join(df_appts, how="inner")
    df_combined.reset_index(inplace=True)
    df_combined["type"] = "Filler"

    # The column names need to be replaced with the appointment mode categories.
    df_combined = df_combined[
        [
            "type",
            "geog_code",
            "geog_ons_code",
            "geog_name",
            "open_practice_count",
            "included_practice_count",
            "total",
            "Face-to-Face",
            "Home Visit",
            "Telephone",
            "Video/Online",
            "Unknown",
        ]
    ]

    # Write the data to sheet
    wb = write_table_to_sheet(wb=wb, sheet_name="Table 3e", table_data=df_combined)
    return wb


def make_and_write_table_4(wb: openpyxl.Workbook) -> openpyxl.Workbook:    
    """
    Writes sheet '4' to the workbook. Loads in data and does some basic organising first. 

    Args:
        wb (openpyxl.Workbook): The workbook loaded from template

    Returns:
        openpyxl.Workbook: The workbook, with the sheet written.
    """    

    df_appts = config.get_appointments_data()
    df_prac_data = config.get_practices_data()

    # Prepare the appointments data
    t4_set = {
        "national_count",
        "by_region",
        "by_stp",
        "by_ccg",
    }
    df_appts = df_appts[df_appts["breakdown"].isin(t4_set)]
    df_appts = df_appts[["geog_code", "geog_ons_code", "geog_name", "appt_count"]]

    # Prepare list size data
    df_prac_data = df_prac_data[["geog_code", "geog_type", "patient_list_size"]]

    # Combine the appointments and practices and list size data
    df_combined = df_prac_data.merge(df_appts, how="inner")
    df_combined.reset_index(inplace=True)

    # Sort and format the data
    custom_dict = {'National': 0, 'Region': 1, 'STP': 2, 'CCG': 3}
    df_combined['rank'] = df_combined['geog_type'].map(custom_dict)

    df_combined.sort_values(by=['rank', 'geog_code'], ascending = [True, True], inplace=True)
    df_combined.drop(columns = ['rank'], inplace=True)

    df_combined["filler_col"] = ""
    column_list = [
        "geog_type",
        "geog_code",
        "geog_ons_code",
        "geog_name",
        "appt_count",
        "filler_col",
        "patient_list_size"
    ] 
    df_combined = df_combined[column_list]

    wb = write_table_to_sheet(wb=wb, sheet_name="Table 4", table_data=df_combined)
    return wb


def make_and_write_table_5(wb: openpyxl.Workbook) -> openpyxl.Workbook:
    """
    Writes sheet 5 to the workbook. Loads in data and does some basic organising first. 

    Args:
        wb (openpyxl.Workbook): The workbook loaded from template

    Returns:
        openpyxl.Workbook: The workbook, with the sheet written.
    """    

    table1_data = config.get_table1_data()
    list_of_months = get_list_of_months()

    # Prepare estimated daily counts from t1_output 
    df_weekday_appts = table1_data[table1_data["breakdown_1"] == "Estimated England total count of appointments by weekday"]
    df_weekday_appts = df_weekday_appts.rename(columns = { 'breakdown_2':'weekday'})
    df_weekday_appts = df_weekday_appts[list_of_months + ['weekday']]
    df_weekday_appts = df_weekday_appts.set_index('weekday').transpose().reset_index().rename(columns = { 'index':'month'})
    df_weekday_appts = df_weekday_appts[['month', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri' ]]
    
    # Prepare coverage percentages from t1_output
    df_coverage = table1_data[table1_data['breakdown_2'] == 'Patient coverage']
    df_coverage = df_coverage.rename(columns = { 'breakdown_2':'patient_coverage'})
    df_coverage  = df_coverage[list_of_months + ['patient_coverage']]
    df_coverage  = df_coverage.set_index('patient_coverage').transpose().reset_index().rename(columns = { 'index':'month'})
    
    # Merge data on month
    df_table_5 = df_weekday_appts.merge(df_coverage, how = 'inner', on = 'month')
    
    wb = write_table_to_sheet(wb=wb, sheet_name="Table 5", table_data=df_table_5)

    return wb


def find_cell_by_tag(
    wb: openpyxl.Workbook, sheet: openpyxl.worksheet, tag: str
) -> Tuple:
    """
    Given a tag and a sheet name, finds the index of that the cell which has that tag

    Args:
        wb (openpyxl.Workbook)
        sheet (openpyxl.worksheet)
        tag (str)

    Returns:
        index of cell containing tag
    """
    max_search = 1000
    ws = wb[sheet]
    iter = 0
    while iter <= max_search:
        for row in ws.iter_rows():
            for cell in row:
                iter += 1
                if cell.value == tag:
                    return openpyxl.utils.cell.coordinate_to_tuple(cell.coordinate)
    return None


# This is a style which can be useful when encountering formatting conflicts. Not currently in use but worth keeping
no_border = openpyxl.styles.borders.Border(
    left=openpyxl.styles.Side(border_style=None),
    top=openpyxl.styles.Side(border_style=None),
    right=openpyxl.styles.Side(border_style=None),
    bottom=openpyxl.styles.Side(border_style=None),
)