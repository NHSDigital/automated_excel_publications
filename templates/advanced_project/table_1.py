# import excel_functions
import openpyxl
from pathlib import Path
# from excel.excel_functions import find_cell_by_tag, find_cell_in_column, get_list_of_months
import pandas as pd
import dateutil.relativedelta
import datetime
from datetime import datetime

import utils
import config

"""
This is a set of functions specifically for dealing with Table 1 in the excel. 
The lists will be used by the 'get_breakdown_dict' to assign each tag a value from dae into a dictionary.
The 'write_table1_month' function goes through each dictionary and uses 'write_single_val' to the template based on the tag used as the key in each dictionary.
The 'write_table_1' function iterates over the given months + columns. 
"""

coverage_list = [
    "<Coverage,Open active practices,count>",
    "<Coverage,Count of practices included,count>",
    "<Coverage,Practice coverage,count>",
    "<Coverage,Registered patients at open active practices,count>",
    "<Coverage,Registered patients at included practices,count>",
    "<Coverage,Patient coverage,count>",
]

working_days_list = [
    "<Working Days,Number of working weekdays,count>",
]

appointment_count_list = [
    "<Appointment Count,Total count of appointments,count>",
    "<Appointment Count,Estimated England total count of appointments,count>",
    "<Appointment Count,Covid Vaccination appointments removed from GP Appointments Data return,count>",
   # "<Appointment Count,Covid Vaccination delivered by practice/PCN,count>",
   # "<Appointment Count,Estimated England total count of appointment including covid vaccinations,count>",
]

appointment_status_list = [
    "<Appointment Status,Attended,count>",
    "<Appointment Status,Did Not Attend,count>",
    "<Appointment Status,Unknown Status,count>",
]

appointment_mode_list = [
    "<Appointment Mode,Face-to-Face,count>",
    "<Appointment Mode,Home Visit,count>",
    "<Appointment Mode,Telephone,count>",
    "<Appointment Mode,Video/Online,count>",
    "<Appointment Mode,Unknown Mode,count>",
]

time_between_list = [
    "<Time between,Same Day,count>",
    "<Time between,1 Day,count>",
    "<Time between,2 to 7 Days,count>",
    "<Time between,8  to 14 Days,count>",
    "<Time between,15  to 21 Days,count>",
    "<Time between,22  to 28 Days,count>",
    "<Time between,More than 28 Days,count>",
    "<Time between,Unknown / Data Quality,count>",
]

hcp_type_list = [
    "<Healthcare Professional,GP,count>",
    "<Healthcare Professional,Other Practice staff,count>",
    "<Healthcare Professional,Unknown HCP,count>",
]

appointment_status_perc_list = [
    "<Appointment Status,Attended,percent>",
    "<Appointment Status,Did Not Attend,percent>",
    "<Appointment Status,Unknown Status,percent>",
]

appointment_mode_perc_list = [
    "<Appointment Mode,Face-to-Face,percent>",
    "<Appointment Mode,Home Visit,percent>",
    "<Appointment Mode,Telephone,percent>",
    "<Appointment Mode,Video/Online,percent>",
    "<Appointment Mode,Unknown Mode,percent>",
]

time_between_perc_list = [
    "<Time between,Same Day,percent>",
    "<Time between,1 Day,percent>",
    "<Time between,2 to 7 Days,percent>",
    "<Time between,8  to 14 Days,percent>",
    "<Time between,15  to 21 Days,percent>",
    "<Time between,22  to 28 Days,percent>",
    "<Time between,More than 28 Days,percent>",
    "<Time between,Unknown / Data Quality,percent>",
]

hcp_type_perc_list = [
    "<Healthcare Professional,GP,percent>",
    "<Healthcare Professional,Other Practice staff,percent>",
    "<Healthcare Professional,Unknown HCP,percent>",
]


def make_and_write_table1(wb: openpyxl.Workbook) -> openpyxl.Workbook:
    """
    Calls write_table1_month() to write the month data to the specified column, iterating through columns + each month in list of months

     Args:
         wb(openpyxl.Workbook): The workbook to edit
         data_path(Path): Path to data folder
         output_path(Path): Path to the output folder
    """
    list_of_months = utils.get_list_of_months()
    column_to_write_to = (
        "C"  # This specifies the first column we want to put a month in
    )
    table1_data = config.get_table1_data()

    for month in list_of_months:
        ws = wb["Table 1"]
        write_table1_month(
            ws=ws, month=month, column=column_to_write_to, table1_data=table1_data
        )
        new_column_number = (
            openpyxl.utils.cell.column_index_from_string(column_to_write_to) + 1
        )
        column_to_write_to = openpyxl.utils.cell.get_column_letter(new_column_number)

    return wb
    

def write_table1_month(
    ws: openpyxl.worksheet, month: str, column: str, table1_data: pd.DataFrame
) -> None:
    """
    Fetches the tag:value dictionary of each breakdown grouping using get_breakdown_dict(), then writes each value using write_single_val()

    Args:
        ws(openpyxl.Worksheet): The worksheet to edit
        month(str): The month with which to fill the month tag in the template
        column(str): The specific column to edit
        table1_data(pd.DataFrame): The table 1 output from dae
    """
    table1_data = table1_data[["breakdown_1", "breakdown_2", "breakdown_3", month]]

    # Month
    month_cell = utils.find_cell_in_column(ws=ws, tag="<month>", column=column)
    ws.cell(row=month_cell[0], column=month_cell[1]).value = month

    # Coverage
    coverage_dict = get_breakdown_dict(month, table1_data, coverage_list)
    for key in coverage_dict:
        write_single_val(ws, column, key, coverage_dict[key])

    # Working Days
    working_days_dict = get_breakdown_dict(month, table1_data, working_days_list)
    for key in working_days_dict:
        write_single_val(ws, column, key, working_days_dict[key])

    # Appointment count
    appointment_count_dict = get_breakdown_dict(
        month, table1_data, appointment_count_list
    )
    for key in appointment_count_dict:
        write_single_val(ws, column, key, appointment_count_dict[key])

    # Appointment status
    appointment_status_dict = get_breakdown_dict(
        month, table1_data, appointment_status_list
    )
    for key in appointment_status_dict:
        write_single_val(ws, column, key, appointment_status_dict[key])

    # Appointment mode
    appointment_mode_dict = get_breakdown_dict(
        month, table1_data, appointment_mode_list
    )
    for key in appointment_mode_dict:
        write_single_val(ws, column, key, appointment_mode_dict[key])

    # Time between booking and appointment
    time_between_dict = get_breakdown_dict(month, table1_data, time_between_list)
    for key in time_between_dict:
        write_single_val(ws, column, key, time_between_dict[key])

    # HCP type
    hcp_type_dict = get_breakdown_dict(month, table1_data, hcp_type_list)
    for key in hcp_type_dict:
        write_single_val(ws, column, key, hcp_type_dict[key])

    # Appointment status perc
    appointment_status_perc_dict = get_breakdown_dict(
        month, table1_data, appointment_status_perc_list
    )
    for key in appointment_status_perc_dict:
        write_single_val(ws, column, key, appointment_status_perc_dict[key])

    # Appointment mode perc
    appointment_mode_perc_dict = get_breakdown_dict(
        month, table1_data, appointment_mode_perc_list
    )
    for key in appointment_mode_perc_dict:
        write_single_val(ws, column, key, appointment_mode_perc_dict[key])

    # Time between booking and appointment perc
    time_between_perc_dict = get_breakdown_dict(
        month, table1_data, time_between_perc_list
    )
    for key in time_between_perc_dict:
        write_single_val(ws, column, key, time_between_perc_dict[key])

    # HCP type perc
    hcp_type_perc_dict = get_breakdown_dict(month, table1_data, hcp_type_perc_list)
    for key in hcp_type_perc_dict:
        write_single_val(ws, column, key, hcp_type_perc_dict[key])

    return None


def write_single_val(ws: openpyxl.worksheet, column: str, tag: str, val: float) -> None:
    """
    Writes a single value into the passed worksheet, identifying the cell using the passed column and tag.

    Args:
        ws(openpyxl.Worksheet): The worksheet to edit
        column(str): The specific column to search
        tag(str): The tag by which to locate the specific cell in the column
        val(float): The value to write to the cell
    """
    loc = utils.find_cell_in_column(ws=ws, tag=tag, column=column)
    ws.cell(row=loc[0], column=loc[1]).value = val

    return None



def get_breakdown_dict(month: str, data: pd.DataFrame, breakdown_list: list) -> dict:
    """
    Takes a list of tags and creates an empty dictionary. It will then break down each tag into it's 3 components and use these to search the Table 1 output of dae for each tag's specific value, which it will assign as the value of the tag in the dictionary.

    Args:
        month(str): The month from which to retrieve data from in the dae output
        data(pd.DataFrame): The table 1 dae output
        breakdown_list(list): A list of tags that specifys specific cells in the excel template
    """
    breakdown_dict = dict.fromkeys(breakdown_list)
    for key in breakdown_dict:
        breakdown_1, breakdown_2, breakdown_3 = key.split(",")
        breakdown_1 = breakdown_1.replace("<", "")
        breakdown_3 = breakdown_3.replace(">", "")
        breakdown_subset = data.loc[
            (
                (data["breakdown_1"] == breakdown_1)
                & (data["breakdown_2"] == breakdown_2)
                & (data["breakdown_3"] == breakdown_3)
            ),
            month,
        ]
        breakdown_dict[key] = breakdown_subset.values[0]

    return breakdown_dict


def make_month_to_write(month: str) -> str:
    """
    Function to add notes to month columns. How we will do this is yet to be agreed.

    Args:
        month(str): The string to be appended with a note
    """
    # TODO
    return month
